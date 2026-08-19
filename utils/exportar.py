"""
Genera el Excel de Papel de Trabajo desde los datos calculados.
v3.2 — Detalle con todas las marcaciones, HE neteadas, período 25-24
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from utils.calculos import calcular_periodo, fmt_dur

MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
            7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

def _fc(h): return PatternFill("solid", fgColor=h)
def _ft(bold=False, sz=10, color="000000"): return Font(bold=bold, size=sz, color=color, name="Calibri")
def _al(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

DARK="1F3864"; BLUE="2E75B6"; LBLUE="D6E4F0"; YEL="FFF2CC"
GREEN="E2EFDA"; LGRY="F2F2F2"; RED="FCE4D6"; WHITE="FFFFFF"

def generar_excel(periodo: str) -> bytes:
    resumen = calcular_periodo(periodo)
    anio, mes = int(periodo[:4]), int(periodo[5:7])
    mes_str = f"{MESES_ES.get(mes, mes)} {anio}"

    # Período real (25 al 24)
    if mes == 1:
        mes_ant, anio_ant = 12, anio - 1
    else:
        mes_ant, anio_ant = mes - 1, anio
    periodo_label = f"25/{mes_ant:02d}/{anio_ant} al 24/{mes:02d}/{anio}"

    wb = Workbook()

    # ══ 1. PANEL GENERAL ══════════════════════════════════════
    ws = wb.active
    ws.title = "Panel General"
    ws.sheet_view.showGridLines = False

    titulo = f"OK ACCESORIOS  ·  PAPEL DE TRABAJO HORARIO  ·  {mes_str.upper()}  ·  {periodo_label}"
    ws.merge_cells("A1:T1")
    c = ws["A1"]; c.value = titulo
    c.font = _ft(True,12,WHITE); c.fill = _fc(DARK); c.alignment = _al()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:T2")
    c = ws["A2"]
    c.value = (f"Período de liquidación: {periodo_label}   |   {len(resumen)} colaboradores   |   "
               f"(*) Horas extra NETAS = HE brutas − minutos de tardanza del mismo día")
    c.font = _ft(False,8,WHITE); c.fill = _fc(BLUE); c.alignment = _al("left")

    # Grupos cabecera fila 3
    grupos = [
        ("A3:C3","IDENTIFICACIÓN",DARK),
        ("D3:I3","ASISTENCIA Y PUNTUALIDAD",BLUE),
        ("J3:K3","HORAS EXTRAS NETAS (*)","375623"),
        ("L3:O3","LICENCIAS Y AUSENCIAS","7D3C00"),
        ("P3:Q3","ADELANTOS / DESC.",DARK),
        ("R3:T3","NOVEDADES","4A235A"),
    ]
    for rng, lbl, bg in grupos:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = lbl; c.font = _ft(True,8,WHITE); c.fill = _fc(bg); c.alignment = _al()

    hdrs = [
        ("A",8,"Legajo"),("B",26,"Apellido y Nombre"),("C",16,"Sector"),
        ("D",7,"Días\nTrab."),("E",7,"Días\nAus."),("F",7,"Feriados"),
        ("G",8,"Home\nOffice"),("H",8,"Cant.\nTardanzas"),("I",10,"Total\nTardanza"),
        ("J",12,"Hs Extra\n50% Netas"),("K",12,"Hs Extra\n100%"),
        ("L",8,"Lic.\nEnf."),("M",8,"Vacac."),("N",8,"ART"),
        ("O",22,"Tipo Ausencia /\nDetalle"),
        ("P",12,"Adelanto $"),("Q",14,"Mercadería $"),
        ("R",12,"Sanción"),("S",14,"Observaciones"),("T",10,"Período"),
    ]
    for col, width, label in hdrs:
        ws.column_dimensions[col].width = width
        c = ws[f"{col}4"]
        c.value = label; c.font = _ft(True,8,WHITE)
        c.fill = _fc("1A3A6B"); c.alignment = _al(wrap=True); c.border = _bd()
    ws.row_dimensions[4].height = 30

    for i, emp in enumerate(resumen):
        row = 5 + i
        bg  = LGRY if i % 2 == 0 else WHITE
        bga = "FFF0E8" if i % 2 == 0 else "FFF8F5"
        bgn = "F5F0FF" if i % 2 == 0 else "FAF7FF"

        partes = []
        if emp["dias_viaje"]: partes.append(f"Viaje ({emp['dias_viaje']}d)")
        if emp["dias_lic"]:   partes.append(f"Lic.Enf ({emp['dias_lic']}d)")
        if emp["dias_vac"]:   partes.append(f"Vacac. ({emp['dias_vac']}d)")
        if emp["dias_art"]:   partes.append(f"ART ({emp['dias_art']}d)")
        tipo_aus = " | ".join(partes)

        vals = [
            (emp["legajo"],bg),(emp["nombre"],bg),(emp["sector"],bg),
            (emp["dias_trab"],bg),(emp["dias_aus"],bg),(emp["dias_feriado"],bg),
            (emp["dias_home"] or "",bg),
            (emp["tard_n"] if emp["tard_n"] else "-",bg),
            (fmt_dur(emp["tard_min"]),bg),
            (fmt_dur(emp["he50"]),bga),(fmt_dur(emp["he100"]),bga),
            (emp["dias_lic"] or "",bga),(emp["dias_vac"] or "",bga),(emp["dias_art"] or "",bga),
            (tipo_aus,bga),
            (emp["adelanto"] if emp["adelanto"] else "",bg),
            (emp["descuento"] if emp["descuento"] else "",bg),
            ("",bgn),("",bgn),(periodo_label,bgn),
        ]
        for ci, (val, bgc) in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = _fc(bgc); c.border = _bd(); c.font = _ft(sz=9)
            c.alignment = _al("left" if ci in (2,3,15,19) else "center")
        ws.row_dimensions[row].height = 15

    # ══ 2. DETALLE POR EMPLEADO (con todas las marcaciones) ════
    ws2 = wb.create_sheet("Detalle por Empleado")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:L1")
    c = ws2["A1"]; c.value = f"OK ACCESORIOS  ·  DETALLE DIARIO  ·  {mes_str.upper()}  ·  {periodo_label}"
    c.font = _ft(True,11,WHITE); c.fill = _fc(DARK); c.alignment = _al()
    ws2.row_dimensions[1].height = 22

    hdrs2 = [
        ("A",8,"Legajo"),("B",22,"Nombre"),("C",14,"Sector"),
        ("D",11,"Fecha"),("E",6,"Día"),("F",20,"Estado"),
        ("G",8,"Entrada"),("H",10,"Ini.\nAlmuerzo"),
        ("I",10,"Fin\nAlmuerzo"),("J",8,"Salida"),
        ("K",10,"Tardanza"),("L",26,"Novedad / Obs."),
    ]
    for col, w, lbl in hdrs2:
        ws2.column_dimensions[col].width = w
        c = ws2[f"{col}2"]; c.value = lbl
        c.font = _ft(True,8,WHITE); c.fill = _fc("1A3A6B")
        c.alignment = _al(wrap=True); c.border = _bd()
    ws2.row_dimensions[2].height = 24

    row2 = 3
    ESTADOS_COLOR = {
        "Trabajó":GREEN,"Tardanza":YEL,"Ausente":RED,
        "Feriado":LBLUE,"Feriado trabajado":"D5E8D4",
        "Domingo":LGRY,"Sábado libre":LGRY,"Sábado HO":"E8F5E9",
    }
    for emp in resumen:
        for dd in emp["detalle"]:
            bg = ESTADOS_COLOR.get(dd["estado"], WHITE)
            if any(k in dd["estado"] for k in ["Viaje","Licencia","Vacaciones","ART","Home office"]):
                bg = "E8F5E9"
            vals = [
                emp["legajo"], emp["nombre"], emp["sector"],
                dd["fecha"].strftime("%d/%m/%Y"), dd["dia"], dd["estado"],
                dd.get("entrada",""), dd.get("ini_almuerzo",""),
                dd.get("fin_almuerzo",""), dd.get("salida",""),
                dd["tardanza"], dd.get("novedad",""),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(row=row2, column=ci, value=val)
                c.fill = _fc(bg); c.border = _bd(); c.font = _ft(sz=8)
                c.alignment = _al("left" if ci in (2,3,6,12) else "center")
            ws2.row_dimensions[row2].height = 13
            row2 += 1

    # ══ 3. HORAS EXTRAS ═══════════════════════════════════════
    ws3 = wb.create_sheet("Horas Extras")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:H1")
    c = ws3["A1"]; c.value = f"OK ACCESORIOS  ·  HORAS EXTRAS NETAS  ·  {mes_str.upper()}  ·  {periodo_label}"
    c.font = _ft(True,11,WHITE); c.fill = _fc(DARK); c.alignment = _al()
    for col, w, lbl in [
        ("A",8,"Legajo"),("B",22,"Nombre"),("C",14,"Sector"),
        ("D",12,"HE 50% Netas"),("E",12,"HE 100%"),("F",12,"TOTAL"),
        ("G",12,"Tardanzas"),("H",22,"Nota"),
    ]:
        ws3.column_dimensions[col].width = w
        c = ws3[f"{col}2"]; c.value = lbl
        c.font = _ft(True,8,WHITE); c.fill = _fc(BLUE)
        c.alignment = _al(); c.border = _bd()
    r = 3
    for emp in resumen:
        if not emp["he50"] and not emp["he100"]: continue
        nota = "HE = HE brutas − tardanzas del día" if emp["tard_min"] else ""
        for ci, val in enumerate([
            emp["legajo"], emp["nombre"], emp["sector"],
            fmt_dur(emp["he50"]), fmt_dur(emp["he100"]),
            fmt_dur(emp["he50"]+emp["he100"]),
            fmt_dur(emp["tard_min"]), nota
        ], 1):
            c = ws3.cell(row=r, column=ci, value=val)
            c.fill = _fc(LGRY if r%2==0 else WHITE); c.border = _bd(); c.font = _ft(sz=9)
            c.alignment = _al("left" if ci in (2,3,8) else "center")
        r += 1

    # ══ 4. TARDANZAS ══════════════════════════════════════════
    ws4 = wb.create_sheet("Tardanzas")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:F1")
    c = ws4["A1"]; c.value = f"OK ACCESORIOS  ·  TARDANZAS  ·  {mes_str.upper()}"
    c.font = _ft(True,11,WHITE); c.fill = _fc(DARK); c.alignment = _al()
    for col, w, lbl in [
        ("A",8,"Legajo"),("B",22,"Nombre"),("C",14,"Sector"),
        ("D",8,"Cant."),("E",12,"Total"),("F",12,"Alerta"),
    ]:
        ws4.column_dimensions[col].width = w
        c = ws4[f"{col}2"]; c.value = lbl
        c.font = _ft(True,8,WHITE); c.fill = _fc(BLUE)
        c.alignment = _al(); c.border = _bd()
    r = 3
    for emp in sorted(resumen, key=lambda x: -x["tard_min"]):
        if not emp["tard_n"]: continue
        alerta = "🔴 ATENCIÓN" if emp["tard_min"] > 60 else "🟡 Informativo"
        for ci, val in enumerate([
            emp["legajo"], emp["nombre"], emp["sector"],
            emp["tard_n"], fmt_dur(emp["tard_min"]), alerta
        ], 1):
            c = ws4.cell(row=r, column=ci, value=val)
            c.fill = _fc(LGRY if r%2==0 else WHITE); c.border = _bd(); c.font = _ft(sz=9)
            c.alignment = _al("left" if ci in (2,3) else "center")
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
